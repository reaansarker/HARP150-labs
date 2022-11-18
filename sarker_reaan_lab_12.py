#!/usr/bin/env python
# coding: utf-8

# workbook- an excel spreadsheet document
# single workbook is saved in a file with .xlsx extension. Each workbook can contain multiple sheets. Each sheet has columns (addressed by letters starting at A) and rows (addressed by numbers starting at 1) A box at a particular column and row is called a cell Each cell can contain a number or text value. The grid of cells with data makes up a sheet.

# In[1]:


import sys
get_ipython().system('conda install --yes --prefix {sys.prefix} openpyxl')


# In[3]:


import openpyxl
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_workbook.title = "Working with OpenPyXL.xlsx"
new_workbook.save(new_workbook.title)
import random
new_sheet['A1'].value = "Ice Cream"
new_sheet['A2'].value = "Vanilla"
new_sheet['A3'].value = "Chocolate"
new_sheet['A4'].value = "Cookies and Cream"
new_sheet['B1'].value = "Quarts"
for i in range(2,5):
    new_sheet['B' + str(i)].value = random.randrange(250)
new_workbook.save(new_workbook.title)
for i in range(2,5):
    if new_sheet['B' + str(i)].value < 100:
        print("You have almost none " + str(new_sheet['A' + str(i)].value))


# In[4]:


get_ipython().system('{sys.executable} -m pip install names')
import names
wb1 = openpyxl.Workbook()
customer_data = wb1.active
customer_data.title = "Customer Data.xlsx"

customer_data['A1'] = "Customer Name"
customer_data['B1'] = "ID"
customer_data['C1'] = "Money Owed"
customer_data['D1'] = "Days Past Due"

for i in range(2, 200):
    customer_data['A' + str(i)] = names.get_full_name()
    customer_data['B' + str(i)] = i
    customer_data['C' + str(i)] = random.randrange(0,5000)
    customer_data['D' + str(i)] = random.randrange(0,365)
wb1.save(customer_data.title)


# In[5]:


import random
wb2 = openpyxl.Workbook()
overdue_customers = wb2.active
overdue_customers.title = "Overdue Customers.xlsx"

overdue_customers['A1'] = "Customer Name"
overdue_customers['B1'] = "ID"
overdue_customers['C1'] = "Money Owed"
overdue_customers['D1'] = "Days Past Due"

for i in range(2, 200):
    customer_data['A' + str(i)] = names.get_full_name()
    customer_data['B' + str(i)] = i
    customer_data['C' + str(i)] = random.randrange(0,5000)
    customer_data['D' + str(i)] = random.randrange(0,366)
wb1.save(customer_data.title)


# In[6]:


import openpyxl
wb3 = openpyxl.load_workbook("Invalid Data.xlsx")
invalid_data = wb3.active

for i in range(2,200):
    if invalid_data['D' + str(i)].value == None or invalid_data['D' + str(i)].value < 0:
        print(invalid_data['A' + str(i)].value + " has invalid Days Past Due value at line " + str(i))


# In[7]:


wb4 = openpyxl.load_workbook("Final Exercise.xlsx")
final_exercise = wb4.active
for i in range(2,200):
    if final_exercise['C' + str(i)].value == None or final_exercise['C' + str(i)].value < 0:
        print(final_exercise['A' + str(i)].value + " has invalid Money Owed value at line " + str(i))
    if final_exercise['D' + str(i)].value == None or final_exercise['D' + str(i)].value < 0:
        print(final_exercise['A' + str(i)].value + " has invalid Days Past Due value at line " + str(i))


# In[11]:


#post lab ex
# https://github.com/jupyter/notebook/issues/5003
# https://zetcode.com/python/openpyxl/


# In[25]:


new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_workbook.title = "sarker_reaan_lab_12.xlsx"
new_workbook.save(new_workbook.title)
import random
new_sheet['A1'].value = "Name" 
new_sheet['B1'].value = "Flavor"
new_sheet['C1'].value = "Number Toppings"
f = open("sarker_reaan_lab_12.xlsx","w",encoding='utf-8')
new_sheet['A2'].value = "Abby"
new_sheet['B2'].value = "Vanilla"
new_sheet['C2'].value = 2

new_sheet['A3'].value = "Coleman"
new_sheet['B3'].value = "Chocolate"
new_sheet['C3'].value = 0

new_sheet['A4'].value = "Sophia"
new_sheet['B4'].value = "Cookies and Cream"
new_sheet['C4'].value = 4

for i in range(2,5):
    if new_sheet['A' + str(i)].value:
        print("What is your name? " + str(new_sheet['A' + str(i)].value) + ". "+ str(new_sheet['A' + str(i)].value) + " ordered a " + str(new_sheet['B' + str(i)].value) + " ice cream with " + str(new_sheet['C' + str(i)].value) + " toppings")
        


# In[ ]:




